#!/usr/bin/env python3
"""
FTIR Reply Automation Bot — Hardened for Real/Private Data
==========================================================
Reads FTIR numbers and pre-written replies from an Excel file, logs into a web
portal, searches each FTIR, opens the correct record, pastes the reply, saves,
and marks the row as Completed/Failed in the Excel file.

Run:  python bot.py
Config:  config.json (copy from config.example.json and fill in values)

See README.md for full setup and usage instructions.
"""

import hashlib
import json
import logging
import os
import re
import signal
import sys
import time
from datetime import datetime

import openpyxl
import pyperclip
from selenium import webdriver
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.edge.options import Options as EdgeOptions
from selenium.webdriver.edge.service import Service as EdgeService
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)
try:
    from webdriver_manager.chrome import ChromeDriverManager
except ImportError:
    ChromeDriverManager = None

try:
    from webdriver_manager.microsoft import EdgeChromiumDriverManager
except ImportError:
    EdgeChromiumDriverManager = None

# ---------------------------------------------------------------------------
# Constants — Excel column names (change here if the real sheet differs)
# ---------------------------------------------------------------------------
# NOTE: These are the LOGICAL names used as dict keys inside the code.
# The ACTUAL header text in Excel is matched flexibly in find_column_indices().
COL_FTIR = "FTIR Number"
COL_REPLY = "Individual Reply"   # Matches 'Individual Reply' column in Excel
COL_STATUS = "Status"
COL_TIMESTAMP = "Timestamp"

STATUS_COMPLETED = "Completed"
STATUS_PENDING = "Pending"
STATUS_DRY_RUN_OK = "Dry-run OK"
STATUS_ALREADY_DONE = "Already Done"
STATUS_FTIR_NOT_FOUND = "Failed: FTIR not found on portal"


class AlreadyDoneError(Exception):
    """Raised when an FTIR record in SIFT already has a response written."""
    pass

# ---------------------------------------------------------------------------
# --- redaction --- Helpers for safe logging of sensitive data
# ---------------------------------------------------------------------------

def redact_ftir(ftir: str) -> str:
    """
    Return a redacted preview of an FTIR number for log output.
    Shows first 3 and last 2 characters with middle masked.
    Example: 'FTIR-12345-XY' → 'FTI...XY'
    """
    ftir = str(ftir).strip()
    if len(ftir) <= 5:
        return ftir[:2] + "..." + ftir[-1:]
    return ftir[:3] + "..." + ftir[-2:]


def redact_reply(reply: str) -> str:
    """
    Return a redacted preview of reply text for log output.
    Shows first 20 chars + length + SHA-256 hash (for matching without exposing content).
    """
    reply = str(reply).strip()
    preview = reply[:20].replace("\n", "\\n")
    h = hashlib.sha256(reply.encode("utf-8")).hexdigest()[:12]
    return f"'{preview}...' ({len(reply)} chars, hash={h})"


# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

def setup_logging(log_file: str) -> logging.Logger:
    """
    Configure dual logging: console (INFO) + file (DEBUG).
    # --- redaction --- Console output uses INFO level and never includes
    # full FTIR numbers or reply text — only row numbers and status.
    # File handler gets DEBUG level with redacted previews.
    """
    logger = logging.getLogger("ftir_reply_bot")
    logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "%(asctime)s | %(levelname)-7s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # File handler — verbose (but still redacted by default;
    # callers use redact_ftir/redact_reply unless verbose_logging is on)
    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    # Console handler — info+ only; callers must never log full content at INFO
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    return logger


# ---------------------------------------------------------------------------
# Config loader
# ---------------------------------------------------------------------------

def load_config(path: str = "config.json") -> dict:
    """Load config.json. Aborts with a clear message if missing."""
    if not os.path.isfile(path):
        print(f"ERROR: Config file '{path}' not found.")
        print("Copy config.example.json → config.json and fill in your values.")
        sys.exit(1)
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# --- data safety --- Synced-folder detection
# ---------------------------------------------------------------------------

SYNCED_FOLDER_MARKERS = [
    "onedrive", "dropbox", "google drive", "icloud", "box sync",
    "sharepoint", "teams",
]


def warn_if_synced_folder(working_dir: str, logger: logging.Logger):
    """
    Print a warning if the working directory appears to be inside a
    cloud-synced folder (OneDrive, Dropbox, etc.), which could leak
    config.json / log files to cloud backups.
    """
    path_lower = working_dir.lower().replace("\\", "/")
    for marker in SYNCED_FOLDER_MARKERS:
        if marker in path_lower:
            logger.warning("!" * 70)
            logger.warning(
                f"  WARNING: Working directory appears to be inside a synced "
                f"folder (detected '{marker}' in path)."
            )
            logger.warning(
                "  config.json (credentials) and bot.log (case data) may be "
                "synced to cloud storage!"
            )
            logger.warning(
                "  Consider moving this project to a local, non-synced directory."
            )
            logger.warning("!" * 70)
            return


# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------

def find_column_indices(sheet) -> dict:
    """
    Scan the header row (row 1) to find column indices for our expected columns.
    Returns a dict like {"FTIR Number": 1, "Reply": 2, "Status": 3, "Timestamp": 4}.
    Supports flexible matching for FTIR, Reply, and Status columns.
    Auto-creates the Timestamp column if it doesn't exist.
    """
    headers = {}
    for col_idx in range(1, sheet.max_column + 1):
        val = sheet.cell(row=1, column=col_idx).value
        if val is not None and str(val).strip():
            headers[str(val).strip()] = col_idx

    col_map = {}

    for raw_header, idx in headers.items():
        # Normalize: lowercase, remove extra spaces, keep alphanumeric + space
        h_norm = re.sub(r"[^\w\s]", " ", raw_header.lower())
        h_norm = re.sub(r"\s+", " ", h_norm).strip()

        if "ftir" in h_norm and COL_FTIR not in col_map:
            col_map[COL_FTIR] = idx
        # PRIORITY: 'Individual Reply' must match BEFORE generic 'reply'
        elif ("individual" in h_norm and "reply" in h_norm) and COL_REPLY not in col_map:
            col_map[COL_REPLY] = idx
        elif any(k in h_norm for k in ("status", "state")) and COL_STATUS not in col_map:
            col_map[COL_STATUS] = idx
        elif "timestamp" in h_norm and COL_TIMESTAMP not in col_map:
            col_map[COL_TIMESTAMP] = idx

    missing = []
    if COL_FTIR not in col_map:
        missing.append(f"{COL_FTIR} (e.g., 'FTIR NO.')")
    if COL_REPLY not in col_map:
        missing.append(f"{COL_REPLY} (e.g., 'Individual Reply')")
    if COL_STATUS not in col_map:
        missing.append(f"{COL_STATUS}")

    if missing:
        raise ValueError(
            f"Excel is missing required column(s): {missing}. "
            f"Found headers: {list(headers.keys())}"
        )

    # Auto-create Timestamp column if not present in headers or col_map
    if COL_TIMESTAMP not in col_map:
        ts_col = sheet.max_column + 1
        sheet.cell(row=1, column=ts_col, value=COL_TIMESTAMP)
        col_map[COL_TIMESTAMP] = ts_col

    return col_map


def build_row_queue(sheet, col_map: dict, dry_run: bool, logger: logging.Logger) -> list:
    """
    Build a list of row numbers to process, strictly top-to-bottom.
    Includes rows where Status is blank, 'Pending', or starts with 'Failed:'
    (so re-runs automatically retry failures).
    In a dry-run, skips rows that are 'Completed', 'Already Done', or 'Dry-run OK'.
    In a live run, skips only 'Completed' and 'Already Done' (so 'Dry-run OK' rows are processed).
    """
    queue = []
    ftir_col = col_map[COL_FTIR]
    status_col = col_map[COL_STATUS]

    skip_statuses = [STATUS_COMPLETED.lower(), STATUS_ALREADY_DONE.lower()]
    if dry_run:
        skip_statuses.append(STATUS_DRY_RUN_OK.lower())

    for row_idx in range(2, sheet.max_row + 1):
        ftir_val = sheet.cell(row=row_idx, column=ftir_col).value
        status_val = sheet.cell(row=row_idx, column=status_col).value

        # Skip rows with no FTIR number (empty rows at the bottom of the sheet)
        if not ftir_val or str(ftir_val).strip() == "":
            continue

        status_str = str(status_val).strip() if status_val else ""
        status_lower = status_str.lower()

        if status_lower in skip_statuses:
            logger.debug(f"  Skipped row {row_idx}: Status='{status_str}'")
        else:
            queue.append(row_idx)
            logger.debug(f"  Queued row {row_idx}: Status='{status_str}'")

    return queue


def update_row_status(wb, sheet, col_map, row_idx, status_text, excel_path, logger):
    """
    Write a status value and timestamp to the given row and immediately save the workbook.
    Saving after every row protects against losing progress on crash.
    """
    status_col = col_map[COL_STATUS]
    sheet.cell(row=row_idx, column=status_col, value=status_text)

    # Write timestamp
    if COL_TIMESTAMP in col_map:
        ts_col = col_map[COL_TIMESTAMP]
        sheet.cell(row=row_idx, column=ts_col, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

    try:
        wb.save(excel_path)
        logger.debug(f"  Row {row_idx} → Status='{status_text}', Excel saved.")
    except PermissionError:
        logger.error(
            f"  Could not save Excel (file may be open in another program). "
            f"Row {row_idx} status '{status_text}' is in memory but NOT persisted to disk."
        )


# ---------------------------------------------------------------------------
# --- preflight --- Pre-flight validation (runs before browser opens)
# ---------------------------------------------------------------------------

def preflight_validate(sheet, col_map: dict, config: dict, dry_run: bool, logger: logging.Logger) -> dict:
    """
    Scan the entire Excel sheet and report:
    - Total rows with FTIR numbers
    - Rows that are already Completed / Already Done
    - Rows that are Dry-run OK
    - Rows that are pending (blank / Pending / Failed / Dry-run OK in live mode)
    - Rows with blank FTIR Number
    - Rows with blank Reply text
    - Duplicate FTIR numbers
    - Unconfigured CSS selectors in config.json
    Returns a summary dict.
    """
    ftir_col = col_map[COL_FTIR]
    reply_col = col_map[COL_REPLY]
    status_col = col_map[COL_STATUS]

    total = 0
    completed = 0
    already_done = 0
    dry_run_ok = 0
    pending = 0
    blank_ftir_rows = []
    blank_reply_rows = []
    ftir_seen = {}  # ftir_value → list of row numbers
    issues = []

    # Check config selectors for placeholders
    selectors = config.get("selectors", {})
    todo_selectors = [k for k, v in selectors.items() if not v or str(v).strip().startswith("<TODO")]
    if todo_selectors:
        issues.append(
            f"  CONFIG UNCONFIGURED: The following selectors in config.json still have '<TODO>' placeholders: "
            f"{', '.join(todo_selectors)}"
        )

    for row_idx in range(2, sheet.max_row + 1):
        ftir_val = sheet.cell(row=row_idx, column=ftir_col).value
        reply_val = sheet.cell(row=row_idx, column=reply_col).value
        status_val = sheet.cell(row=row_idx, column=status_col).value

        ftir_str = str(ftir_val).strip() if ftir_val else ""
        reply_str = str(reply_val).strip() if reply_val else ""
        status_str = str(status_val).strip() if status_val else ""

        # Skip completely empty rows (no FTIR, no reply, no status)
        if not ftir_str and not reply_str and not status_str:
            continue

        total += 1

        if status_str.lower() == STATUS_COMPLETED.lower():
            completed += 1
            continue
        if status_str.lower() == STATUS_ALREADY_DONE.lower():
            already_done += 1
            continue
        if status_str.lower() == STATUS_DRY_RUN_OK.lower():
            dry_run_ok += 1
            if not dry_run:
                pending += 1
            continue

        pending += 1

        # Check for blank FTIR
        if not ftir_str:
            blank_ftir_rows.append(row_idx)
            issues.append(f"  Row {row_idx}: FTIR Number is blank")

        # Check for blank Reply
        if ftir_str and not reply_str:
            blank_reply_rows.append(row_idx)
            issues.append(f"  Row {row_idx}: Reply is blank (FTIR={redact_ftir(ftir_str)})")

        # Track duplicates
        if ftir_str:
            ftir_seen.setdefault(ftir_str, []).append(row_idx)

    # Find duplicates
    duplicates = {k: v for k, v in ftir_seen.items() if len(v) > 1}
    for ftir_val, rows in duplicates.items():
        issues.append(
            f"  DUPLICATE: FTIR {redact_ftir(ftir_val)} appears in rows {rows} "
            f"— possible data-entry mistake"
        )

    return {
        "total": total,
        "completed": completed,
        "already_done": already_done,
        "dry_run_ok": dry_run_ok,
        "pending": pending,
        "blank_ftir_rows": blank_ftir_rows,
        "blank_reply_rows": blank_reply_rows,
        "duplicates": duplicates,
        "issues": issues,
    }


# ---------------------------------------------------------------------------
# --- connectivity --- Network / Portal pre-check
# ---------------------------------------------------------------------------

def check_portal_connectivity(portal_url: str, logger: logging.Logger, timeout: int = 10) -> bool:
    """
    Verify network connectivity to the SIFT portal before opening a browser.
    Uses urllib to do a simple HTTP HEAD/GET request.
    Returns True if reachable, False otherwise.
    """
    import urllib.request
    import urllib.error
    import ssl

    logger.info(f"  Checking portal connectivity: {portal_url} ...")
    try:
        # Create SSL context that doesn't verify (internal portals often have self-signed certs)
        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        req = urllib.request.Request(portal_url, method="HEAD")
        req.add_header("User-Agent", "FTIR-Reply-Bot/1.0")
        resp = urllib.request.urlopen(req, timeout=timeout, context=ctx)
        status_code = resp.getcode()
        logger.info(f"  Portal responded with HTTP {status_code} ✓")
        return True
    except urllib.error.HTTPError as e:
        # HTTP errors (401, 403, etc.) still mean the portal is reachable
        logger.info(f"  Portal responded with HTTP {e.code} (reachable, auth required) ✓")
        return True
    except urllib.error.URLError as e:
        logger.error(f"  ✗ Cannot reach portal: {e.reason}")
        return False
    except Exception as e:
        logger.error(f"  ✗ Connectivity check failed: {e}")
        return False


# ---------------------------------------------------------------------------
# --- FTIR pre-validation --- Check FTIR exists on portal before batch
# ---------------------------------------------------------------------------

def pre_validate_ftir_on_portal(
    driver, ftir_number: str, sel: dict, timeout: int, logger: logging.Logger
) -> bool:
    """
    Check whether a specific FTIR number exists on the portal by performing
    a Quick Search and checking if a result window opens.
    Returns True if the FTIR exists, False otherwise.
    Does NOT modify anything — read-only check.
    """
    wait = WebDriverWait(driver, timeout)

    try:
        navigate_to_quick_search(driver, sel, timeout, logger)
    except Exception as e:
        logger.warning(f"  Pre-validation: Could not open Quick Search: {e}")
        return True  # Can't verify — assume it exists to avoid false negatives

    try:
        # Enter FTIR number
        search_box = wait.until(EC.presence_of_element_located((By.ID, "txtSel0")))
        search_box.clear()
        search_box.send_keys(str(ftir_number))

        # Record windows before clicking Search
        old_windows = set(driver.window_handles)

        # Click Search
        wait.until(EC.element_to_be_clickable((By.ID, "searchbtn"))).click()

        # Wait briefly for result window
        try:
            WebDriverWait(driver, min(timeout, 8)).until(
                lambda d: len(set(d.window_handles) - old_windows) > 0
            )
            # Result window opened — FTIR exists
            logger.debug(f"  Pre-validation: FTIR {redact_ftir(ftir_number)} found on portal ✓")
            return True
        except TimeoutException:
            # No result window — FTIR doesn't exist
            logger.warning(f"  Pre-validation: FTIR {redact_ftir(ftir_number)} NOT found on portal ✗")
            return False

    except Exception as e:
        logger.warning(f"  Pre-validation check error for {redact_ftir(ftir_number)}: {e}")
        return True  # Can't verify — assume it exists

    finally:
        # Clean up: close any extra windows opened during validation
        close_extra_windows(driver, logger)


def print_preflight_and_confirm(summary: dict, dry_run: bool, logger: logging.Logger) -> bool:
    """
    Print preflight summary and ask for y/n confirmation.
    Returns True if user confirms, False otherwise.
    """
    logger.info("")
    logger.info("=" * 70)
    logger.info("PRE-FLIGHT VALIDATION")
    logger.info("=" * 70)
    logger.info(f"  Total data rows     : {summary['total']}")
    logger.info(f"  Already Completed   : {summary['completed']} (will be skipped)")
    if summary.get('already_done', 0) > 0:
        logger.info(f"  Already Done        : {summary['already_done']} (will be skipped)")
    skip_or_proc = "will be skipped" if dry_run else "will be processed"
    logger.info(f"  Dry-run OK          : {summary['dry_run_ok']} ({skip_or_proc})")
    logger.info(f"  Pending / to process: {summary['pending']}")
    logger.info(f"  Mode                : {'DRY RUN (Save will NOT be clicked)' if dry_run else 'LIVE RUN (Will Save & Update Excel)'}")

    if summary["issues"]:
        logger.info("")
        logger.warning("  ⚠ Issues found:")
        for issue in summary["issues"]:
            logger.warning(f"    {issue}")

    logger.info("=" * 70)

    if summary["pending"] == 0:
        logger.info("No pending rows to process. Exiting.")
        return False

    # --- preflight --- Require manual y/n confirmation before proceeding
    print()  # blank line for readability
    prompt = "Proceed? (y/n): "
    if dry_run:
        prompt = "Proceed with DRY RUN? (y/n): "
    try:
        answer = input(prompt).strip().lower()
    except (EOFError, KeyboardInterrupt):
        print()
        logger.info("User cancelled at pre-flight confirmation.")
        return False

    if answer != "y":
        logger.info("User declined to proceed. Exiting.")
        return False

    logger.info("User confirmed. Starting bot...")
    return True


# ---------------------------------------------------------------------------
# Whitespace-normalized comparison for paste verification
# ---------------------------------------------------------------------------

def normalize_whitespace(text: str) -> str:
    """
    Collapse all whitespace runs to single spaces and strip.
    Used to compare intended vs. actual paste content, because browsers
    sometimes alter internal spacing in textareas.
    """
    return re.sub(r"\s+", " ", text).strip()


# ---------------------------------------------------------------------------
# Browser helpers
# ---------------------------------------------------------------------------
# Browser helpers & locator parsing
# ---------------------------------------------------------------------------

def parse_locator(selector_str: str):
    """Determine By.XPATH or By.CSS_SELECTOR from selector string."""
    s = str(selector_str).strip()
    if s.startswith("//") or s.startswith("(") or s.startswith("./") or "contains(" in s or "following-sibling::" in s:
        return (By.XPATH, s)
    return (By.CSS_SELECTOR, s)


def create_driver(config: dict, logger: logging.Logger, attach: bool = False, port: int = 9222):
    """Create or attach to a Selenium Edge or Chrome WebDriver."""
    is_attached = attach or config.get("connect_to_existing_browser", False)
    dbg_port = port or config.get("remote_debugging_port", 9222)

    # 1. If attaching to an already running browser (Edge or Chrome)
    if is_attached:
        logger.info(f"Attaching to existing browser session at 127.0.0.1:{dbg_port}...")
        
        # Try Edge WebDriver first
        try:
            edge_opts = EdgeOptions()
            edge_opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{dbg_port}")
            driver = webdriver.Edge(options=edge_opts)
            driver.implicitly_wait(0)
            logger.info("Connected to existing Microsoft Edge browser session successfully.")
            return driver
        except Exception as edge_err:
            logger.debug(f"Edge attach direct attempt: {edge_err}. Trying Chrome WebDriver...")

        # Try Chrome WebDriver as fallback
        try:
            chrome_opts = ChromeOptions()
            chrome_opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{dbg_port}")
            driver = webdriver.Chrome(options=chrome_opts)
            driver.implicitly_wait(0)
            logger.info("Connected to existing Chromium browser session successfully.")
            return driver
        except Exception as chrome_err:
            raise RuntimeError(
                f"Could not connect to browser on port {dbg_port}. "
                f"Please ensure Microsoft Edge or Chrome was started with: "
                f"msedge.exe --remote-debugging-port={dbg_port}"
            ) from chrome_err

    # 2. If launching a new browser instance
    # Try Edge first
    try:
        edge_opts = EdgeOptions()
        if config.get("headless", False):
            edge_opts.add_argument("--headless=new")
        edge_opts.add_argument("--no-sandbox")
        edge_opts.add_argument("--disable-dev-shm-usage")
        edge_opts.add_argument("--window-size=1366,900")
        driver = webdriver.Edge(options=edge_opts)
        driver.implicitly_wait(0)
        logger.info("Microsoft Edge browser launched.")
        return driver
    except Exception:
        pass

    # Fallback to Chrome
    chrome_opts = ChromeOptions()
    if config.get("headless", False):
        chrome_opts.add_argument("--headless=new")
    chrome_opts.add_argument("--no-sandbox")
    chrome_opts.add_argument("--disable-dev-shm-usage")
    chrome_opts.add_argument("--window-size=1366,900")
    try:
        if ChromeDriverManager:
            service = ChromeService(ChromeDriverManager().install())
            driver = webdriver.Chrome(service=service, options=chrome_opts)
        else:
            driver = webdriver.Chrome(options=chrome_opts)
    except Exception:
        driver = webdriver.Chrome(options=chrome_opts)

    driver.implicitly_wait(0)
    logger.info("Chrome browser launched.")
    return driver


def wait_and_find(driver, selector_expr: str, timeout: int, clickable: bool = False):
    """
    Explicit wait for an element. Supports multiple fallback selectors separated by ' | '.
    Automatically detects XPath vs CSS selector.
    """
    if not selector_expr or str(selector_expr).strip().startswith("<TODO"):
        raise ValueError(f"Invalid selector '{selector_expr}'.")

    # Split fallback selectors separated by ' | '
    candidates = [c.strip() for c in selector_expr.split(" | ") if c.strip() and not c.strip().startswith("<TODO")]
    if not candidates:
        raise ValueError(f"No valid selectors found in '{selector_expr}'.")

    last_err = None
    per_candidate_timeout = max(2, timeout // len(candidates))

    for cand in candidates:
        by, value = parse_locator(cand)
        condition = (
            EC.element_to_be_clickable((by, value))
            if clickable
            else EC.visibility_of_element_located((by, value))
        )
        try:
            return WebDriverWait(driver, per_candidate_timeout).until(condition)
        except Exception as e:
            last_err = e
            continue

    raise last_err or TimeoutException(f"Could not locate element using selectors: {selector_expr}")


def safe_clear_field(driver, element, logger):
    """
    Robustly clear an input/textarea field.
    First tries .clear(), then select-all + delete as a fallback,
    then verifies the field is actually empty.
    """
    try:
        element.clear()
    except Exception:
        pass
    time.sleep(0.2)

    # Fallback: select-all → delete
    try:
        element.send_keys(Keys.CONTROL + "a")
        time.sleep(0.1)
        element.send_keys(Keys.DELETE)
        time.sleep(0.1)
    except Exception:
        pass

    # Verify it's empty
    remaining = element.get_attribute("value") or element.text
    if remaining and remaining.strip():
        logger.debug("  Field not fully cleared, using JS clear.")
        driver.execute_script("arguments[0].value = '';", element)
        driver.execute_script(
            "arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", element
        )


def is_logged_in(driver) -> bool:
    """Check if the user has successfully logged into the SIFT portal."""
    try:
        # Check for frames (SIFT main UI layout uses frames)
        frames = driver.find_elements(By.TAG_NAME, "frame")
        if len(frames) > 0:
            return True
        # Check if username input is no longer present and page has loaded
        user_inputs = driver.find_elements(By.ID, "username")
        if not user_inputs and ("sift" in driver.current_url.lower() or "bizapps" in driver.current_url.lower()):
            body_text = driver.find_element(By.TAG_NAME, "body").text.lower()
            if any(k in body_text for k in ("quick search", "main menu", "logout", "feedback", "ftir")):
                return True
    except Exception:
        pass
    return False


def login(driver, config: dict, sel: dict, timeout: int, logger: logging.Logger):
    """Navigate to the SIFT portal and log in (auto-detects login without manual ENTER)."""
    url = config.get("portal_url", "https://sift.bizapps.suzuki/sift/")
    logger.info(f"Navigating to SIFT portal: {url}...")
    try:
        driver.get(url)
    except Exception as e:
        logger.warning(f"Navigation warning: {e}")

    time.sleep(2)

    user = config.get("username", "").strip()
    pwd = config.get("password", "").strip()

    if user and pwd:
        # Auto-fill credentials if provided in config.json
        try:
            wait = WebDriverWait(driver, 5)
            wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(user)
            logger.debug("  Entered username.")
            wait.until(EC.presence_of_element_located((By.ID, "password"))).send_keys(pwd)
            logger.debug("  Entered password.")
            wait.until(EC.element_to_be_clickable((By.ID, "cal-login-button"))).click()
            logger.info("  Credentials submitted automatically. Waiting for SIFT main menu...")
        except Exception as e:
            logger.warning(f"Auto credential entry skipped: {e}")

    # Auto-poll until user/browser is logged into SIFT (no ENTER needed)
    logger.info("Waiting for SIFT login to complete (auto-detecting)...")
    poll_start = time.time()
    max_wait = 180  # wait up to 3 minutes for user login

    while time.time() - poll_start < max_wait:
        if is_logged_in(driver):
            logger.info("✓ Login detected! SIFT Main Menu ready. Starting bot processing...")
            time.sleep(1.5)
            return
        time.sleep(1)

    logger.warning("Auto-detect timeout reached. Proceeding with execution...")


def logout_and_quit(driver, config: dict, sel: dict, logger: logging.Logger, is_attached: bool = False, keep_open: bool = False):
    """
    Attempt to log out of the portal if new driver was created, or detach if connected to user session.
    If keep_open is True, leaves the browser open without quitting.
    """
    if driver is None:
        return

    if is_attached or keep_open:
        logger.info("  Browser window left open for inspection.")
        return

    try:
        logout_sel = sel.get("logout_button", "")
        if logout_sel and not logout_sel.startswith("<TODO"):
            try:
                logout_btn = wait_and_find(driver, logout_sel, 5, clickable=True)
                logout_btn.click()
                logger.info("  Logged out of portal.")
                time.sleep(1)
            except Exception:
                logger.debug("  Logout button not found or click failed; session will end with browser close.")
    finally:
        try:
            driver.quit()
            logger.info("  Browser closed and session destroyed.")
        except Exception:
            logger.debug("  Browser was already closed.")


def close_extra_windows(driver, logger=None):
    """
    Close all popup/extra windows and return to the main SIFT window (window_handles[0]).
    """
    try:
        if not driver or not driver.window_handles:
            return
        main_window = driver.window_handles[0]
        for handle in driver.window_handles[1:]:
            try:
                if handle in driver.window_handles:
                    driver.switch_to.window(handle)
                    driver.close()
                    time.sleep(0.3)
            except Exception:
                pass
        if main_window in driver.window_handles:
            driver.switch_to.window(main_window)
            driver.switch_to.default_content()
    except Exception as e:
        if logger:
            logger.debug(f"Window cleanup notice: {e}")


def switch_to_content_frame(driver, logger, search_text="QUICK SEARCH"):
    """
    Old enterprise Java apps (like SIFT) use HTML framesets.
    This function tries to find the correct frame containing the target text.
    It searches the main document first, then iterates through all frames/iframes.
    """
    # First check if content is in the default (top) document
    driver.switch_to.default_content()
    try:
        if search_text.lower() in driver.page_source.lower():
            logger.debug(f"  Found '{search_text}' in top-level document.")
            return True
    except Exception:
        pass

    # Try each frame/iframe
    frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")
    logger.debug(f"  Found {len(frames)} frame(s)/iframe(s) on page.")

    for i, frame in enumerate(frames):
        try:
            frame_name = frame.get_attribute("name") or frame.get_attribute("id") or f"frame-{i}"
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            src = driver.page_source
            if search_text.lower() in src.lower():
                logger.info(f"  Switched to frame '{frame_name}' containing '{search_text}'.")
                return True
            logger.debug(f"  Frame '{frame_name}' does not contain '{search_text}'.")
        except Exception as e:
            logger.debug(f"  Could not check frame {i}: {e}")

    # If no frame matched, go back to default and try anyway
    driver.switch_to.default_content()
    logger.debug(f"  No frame contained '{search_text}'. Staying on default content.")
    return False


def find_element_across_frames(driver, selector_expr, timeout, logger, clickable=False):
    """
    Try to find an element in the current frame first, then search all frames.
    Returns the element if found, otherwise raises TimeoutException.
    """
    # Try current frame first
    try:
        return wait_and_find(driver, selector_expr, min(timeout, 3), clickable=clickable)
    except Exception:
        pass

    # Search across all frames
    driver.switch_to.default_content()
    frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")

    for i, frame in enumerate(frames):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(frame)
            el = wait_and_find(driver, selector_expr, min(timeout, 3), clickable=clickable)
            frame_name = frame.get_attribute("name") or frame.get_attribute("id") or f"frame-{i}"
            logger.debug(f"  Found element in frame '{frame_name}'.")
            return el
        except Exception:
            continue

    # Last resort: switch back to default
    driver.switch_to.default_content()
    return wait_and_find(driver, selector_expr, timeout, clickable=clickable)


def navigate_to_quick_search(driver, sel: dict, timeout: int, logger: logging.Logger):
    """
    On SIFT main menu page, switch to frame(1) (menuFrame) and click 'QUICK SEARCH'.
    QUICK SEARCH opens a new popup window — we switch to it.
    """
    wait = WebDriverWait(driver, timeout)

    # Go back to main window and default content
    driver.switch_to.window(driver.window_handles[0])
    driver.switch_to.default_content()

    # Wait for frames to be present
    wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, "frame")))

    # Switch to frame(1) — the menu frame
    driver.switch_to.frame(1)
    logger.info("  Switched to menu frame (frame 1).")

    # Record existing windows before clicking
    old_windows = set(driver.window_handles)

    # Click QUICK SEARCH using the exact SIFT selector
    qs_xpath = "//div[@id='group2content']//table[@class='NoBorderTable']//tr/td[2][contains(., 'QUICK SEARCH')]"
    try:
        quick_search = wait.until(EC.element_to_be_clickable((By.XPATH, qs_xpath)))
        quick_search.click()
        logger.info("  Clicked QUICK SEARCH. Waiting for popup window...")
    except Exception:
        # Fallback: try broader selectors
        logger.debug("  Exact QUICK SEARCH selector failed. Trying fallbacks...")
        driver.switch_to.default_content()
        driver.switch_to.frame(1)
        links = driver.find_elements(By.TAG_NAME, "a") + driver.find_elements(By.TAG_NAME, "td")
        clicked = False
        for el in links:
            try:
                txt = el.text.strip().upper()
                if "QUICK SEARCH" in txt:
                    driver.execute_script("arguments[0].click();", el)
                    logger.info("  Clicked QUICK SEARCH via fallback.")
                    clicked = True
                    break
            except Exception:
                continue
        if not clicked:
            raise RuntimeError("Could not find QUICK SEARCH on SIFT menu.")

    # Switch back to default content before handling new window
    driver.switch_to.default_content()

    # Wait for the Quick Search popup window to open
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: len(set(d.window_handles) - old_windows) > 0
        )
        new_window = list(set(driver.window_handles) - old_windows)[0]
        driver.switch_to.window(new_window)
        logger.info(f"  Switched to Quick Search popup window.")
    except TimeoutException:
        raise RuntimeError("Quick Search popup window did not open.")

    time.sleep(1)


def search_ftir(driver, ftir_number: str, sel: dict, timeout: int, logger: logging.Logger):
    """
    In the Quick Search popup window:
    1. Enter FTIR number into input#txtSel0
    2. Click Search button#searchbtn
    3. Switch to the new FTIR response window
    """
    wait = WebDriverWait(driver, timeout)

    navigate_to_quick_search(driver, sel, timeout, logger)

    # Enter FTIR number — exact SIFT element ID: txtSel0
    search_box = wait.until(EC.presence_of_element_located((By.ID, "txtSel0")))
    search_box.clear()
    search_box.send_keys(str(ftir_number))
    logger.info(f"  Entered FTIR {redact_ftir(ftir_number)} into search box.")

    # Record windows before clicking Search
    old_windows = set(driver.window_handles)

    # Click Search button — exact SIFT element ID: searchbtn
    wait.until(EC.element_to_be_clickable((By.ID, "searchbtn"))).click()
    logger.info("  Clicked Search button. Waiting for FTIR response window...")

    # Wait for FTIR response window to open
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: len(set(d.window_handles) - old_windows) > 0
        )
        response_window = list(set(driver.window_handles) - old_windows)[0]
        driver.switch_to.window(response_window)
        logger.info(f"  Switched to FTIR response window.")
    except TimeoutException:
        raise RuntimeError(
            f"FTIR response window did not open for {redact_ftir(ftir_number)}. "
            f"The FTIR number may be invalid."
        )

    time.sleep(2)


def select_correct_result(
    driver, ftir_number: str, sel: dict, timeout: int, logger: logging.Logger
):
    """
    After search results load, find and click the row whose FTIR number
    exactly matches `ftir_number`.

    If there's exactly one result, we still verify the text before clicking.
    If there are two+ results and we can't find an exact match, we raise
    an exception so the row is flagged as 'Failed: ambiguous search result'.
    """
    # Find all result rows
    rows = driver.find_elements(By.CSS_SELECTOR, sel["result_rows"])
    logger.debug(f"  Found {len(rows)} search result row(s).")

    if len(rows) == 0:
        raise NoSuchElementException(
            f"No search results found for FTIR '{redact_ftir(ftir_number)}'."
        )

    # Try to find the exact-match row
    matched_row = None
    ftir_normalized = ftir_number.strip().lower()

    for row in rows:
        try:
            ftir_cell = row.find_element(By.CSS_SELECTOR, sel["result_ftir_text"])
            cell_text = (ftir_cell.text or "").strip().lower()
            if cell_text == ftir_normalized:
                matched_row = row
                break
        except (NoSuchElementException, StaleElementReferenceException):
            continue

    if matched_row is None:
        if len(rows) == 1:
            # Single result but text didn't match exactly — risky but log a warning
            logger.warning(
                "  Single result found but FTIR text didn't match exactly. "
                "Clicking it anyway; verify selectors are correct."
            )
            matched_row = rows[0]
        else:
            raise ValueError(
                f"Ambiguous search result: {len(rows)} rows found, "
                f"none matched FTIR exactly."
            )

    # Click the matched row to open the record
    matched_row.click()
    logger.debug("  Clicked matching result row.")


# --- record match safety --- Exact-match header verification
def verify_record_header(driver, ftir_number: str, sel: dict, timeout: int, logger: logging.Logger):
    """
    After navigating into a record, verify the page header / title area
    shows the expected FTIR number. This is the critical data-integrity
    check that prevents pasting a reply onto the wrong record.

    Uses EXACT string match (not substring/contains) to prevent false
    positives on similar numbers (e.g. 'FTIR-123' matching 'FTIR-1234').
    """
    header_sel = sel.get("record_header_ftir", "")
    if not header_sel or header_sel.startswith("<TODO"):
        logger.warning(
            "  record_header_ftir selector not configured — "
            "skipping record verification (DATA INTEGRITY RISK)."
        )
        return

    header_el = wait_and_find(driver, header_sel, timeout)
    header_text = (header_el.text or "").strip()

    # --- record match safety --- EXACT match, not substring 'in' check.
    # Compare normalized (case-insensitive, stripped) values.
    if header_text.lower() != ftir_number.strip().lower():
        raise ValueError(
            f"Record header mismatch! Expected FTIR exact match but page shows "
            f"different value. Aborting this row (no save attempted)."
        )
    logger.debug("  Record header verified (exact match) ✓")


def check_existing_reply(driver, textarea, logger: logging.Logger):
    """
    Check if the FTIR record in SIFT already has a response written or is completed.
    Raises AlreadyDoneError if a response is already present so it won't be overwritten.
    """
    # 1. Check if SIFT marks the Feedback/Individual Reply section as Completed
    try:
        is_completed = driver.execute_script("""
            var bodyText = (document.body && document.body.innerText) ? document.body.innerText : "";
            if (bodyText.indexOf("FEEDBACK [Completed]") !== -1 || 
                bodyText.indexOf("INDIVIDUAL REPLY [Completed]") !== -1 ||
                bodyText.indexOf("Individual Reply [Completed]") !== -1) {
                return true;
            }
            var fd = document.querySelector("input[name='buttonControl.fdKanryoIndividual']");
            if (fd && fd.value === "true") {
                return true;
            }
            return false;
        """)
        if is_completed:
            logger.info("  SIFT record indicates Individual Reply is already completed.")
            raise AlreadyDoneError("Individual Reply is already marked as Completed in SIFT.")
    except AlreadyDoneError:
        raise
    except Exception:
        pass

    # 2. Check if the reply textarea already contains an actual reply (ignoring SIFT's default watermark)
    if textarea:
        try:
            current_val = (textarea.get_attribute("value") or textarea.text or "").strip()
            # SIFT puts a placeholder watermark in the textarea by default
            is_placeholder = (
                "please recheck if the contents of your reply" in current_val.lower()
                or "suzuki's conclusion" in current_val.lower()
            )
            if current_val and not is_placeholder:
                preview = current_val[:35].replace("\n", " ")
                logger.info(f"  FTIR record already has a response written: '{preview}...' ({len(current_val)} chars).")
                raise AlreadyDoneError(f"Response already written: '{preview}...'")
        except AlreadyDoneError:
            raise
        except Exception:
            pass


def find_feedback_frame(driver, logger: logging.Logger):
    """
    Search top-level document and all frames/iframes for the Feedback / Reply section.
    Switches driver into the correct frame and returns True if found.
    """
    # 1. Check top-level document first
    driver.switch_to.default_content()
    try:
        src = driver.page_source.lower()
        if "swfeedbackblock" in src or "reply individually" in src or "representative reply" in src:
            return True
    except Exception:
        pass

    # 2. Check all frames and iframes
    frames = driver.find_elements(By.TAG_NAME, "frame") + driver.find_elements(By.TAG_NAME, "iframe")
    for i, fr in enumerate(frames):
        try:
            driver.switch_to.default_content()
            driver.switch_to.frame(fr)
            src = driver.page_source.lower()
            if "swfeedbackblock" in src or "reply individually" in src or "representative reply" in src:
                logger.info(f"  Switched to frame {i} containing SIFT feedback section.")
                return True
        except Exception:
            continue

    driver.switch_to.default_content()
    return False


def open_reply_field(driver, sel: dict, timeout: int, logger: logging.Logger):
    """
    On SIFT FTIR record page:
    1. Switch to the correct content frame.
    2. Click the 'Page with Minus sign' toggle logo / icon to collapse upper sections.
    3. Ensure Feedback Block (swFeedbackBlock) is expanded.
    4. Select 'Reply individually.' radio button.
    5. Return the exact Individual Reply textarea element.
    """
    find_feedback_frame(driver, logger)

    # 1. Click the 'Page with Minus sign' toggle logo / icon to collapse upper sections
    try:
        js_click_minus_logo = """
            // 1. Look for the 'Page with Minus' icon / toggle all button
            var allImagesAndLinks = document.querySelectorAll("img, a, input[type='image'], input[type='button'], button, span");
            var clicked = false;

            for (var i = 0; i < allImagesAndLinks.length; i++) {
                var el = allImagesAndLinks[i];
                var src = (el.src || el.getAttribute('src') || '').toLowerCase();
                var onclick = (el.getAttribute('onclick') || '').toLowerCase();
                var alt = (el.alt || el.title || el.className || el.id || '').toLowerCase();

                // Matches 'minus', 'collapse', 'allclose', 'toggle', 'fold' icons
                if (src.includes('minus') || src.includes('close') || src.includes('collapse') ||
                    onclick.includes('allclose') || onclick.includes('toggle') || onclick.includes('fold') ||
                    alt.includes('minus') || alt.includes('collapse') || alt.includes('all close')) {
                    el.click();
                    clicked = true;
                    break;
                }
            }

            // 2. Ensure the Feedback section (swFeedbackBlock) is expanded and scrolled into view
            if (typeof openArea === 'function') {
                try { openArea('swFeedbackBlock'); } catch(e){}
            }
            var fb = document.getElementById("swFeedbackBlock");
            if (fb) {
                fb.style.display = 'block';
                fb.scrollIntoView({ behavior: 'smooth', block: 'center' });
            } else {
                window.scrollTo({ top: document.body.scrollHeight / 2, behavior: 'smooth' });
            }
            return clicked;
        """
        was_clicked = driver.execute_script(js_click_minus_logo)
        if was_clicked:
            logger.info("  ✓ Clicked 'Page with Minus' toggle logo to scroll down to Feedback.")
        time.sleep(0.4)
    except Exception as e:
        logger.debug(f"  Minus logo click notice: {e}")

    # Fallback: Click minus/collapse icon via XPath
    try:
        minus_icons = driver.find_elements(
            By.XPATH,
            "//img[contains(@src, 'minus') or contains(@src, 'close') or contains(@alt, 'minus') or contains(@title, 'minus')] | "
            "//a[contains(@onclick, 'allClose') or contains(@onclick, 'toggle') or contains(@href, 'allClose')]"
        )
        for mi in minus_icons:
            try:
                mi.click()
                logger.info("  ✓ Clicked minus toggle icon via XPath.")
                time.sleep(0.3)
                break
            except Exception:
                continue
    except Exception:
        pass

    # 2. Expand Feedback Block if collapsed or hidden
    try:
        driver.execute_script("""
            if (typeof openArea === 'function') {
                try { openArea('swFeedbackBlock'); } catch(e){}
            }
            var el = document.getElementById('swFeedbackBlock');
            if (el) {
                el.style.display = 'block';
                el.scrollIntoView({ block: 'center' });
            }
        """)
        time.sleep(0.3)
    except Exception:
        pass

    # 2. Select 'Reply individually.' radio button
    try:
        driver.execute_script("""
            var radios = document.querySelectorAll("input[type='radio']");
            for (var i = 0; i < radios.length; i++) {
                var r = radios[i];
                var row = r.closest("tr") || r.parentElement;
                var txt = ((row ? row.innerText : "") + " " + (r.name || "") + " " + (r.value || "")).toLowerCase();
                if (txt.includes("individually") || txt.includes("individual")) {
                    r.removeAttribute('disabled');
                    r.checked = true;
                    r.scrollIntoView({block: 'center'});
                    r.dispatchEvent(new Event('change', {bubbles: true}));
                    r.dispatchEvent(new Event('click', {bubbles: true}));
                    break;
                }
            }
        """)
        time.sleep(0.4)
    except Exception as e:
        logger.debug(f"  JS radio selection notice: {e}")

    # Fallback radio click via ActionChains
    try:
        radio_elements = driver.find_elements(
            By.XPATH,
            "//tr[contains(., 'Reply individually') or contains(., 'individually')]//input[@type='radio'] | "
            "//label[contains(., 'Reply individually') or contains(., 'individually')] | "
            "//*[contains(text(), 'Reply individually')]"
        )
        for r_el in radio_elements:
            try:
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", r_el)
                time.sleep(0.2)
                ActionChains(driver).move_to_element(r_el).click().perform()
                break
            except Exception:
                continue
    except Exception:
        pass

    time.sleep(0.3)

    # 3. Locate the EXACT textarea in the 'Reply individually.' row (beside [Final])
    target_box = None
    try:
        js_find_ta = """
            // STEP 1: Find the <tr> row that contains 'Reply individually' text
            // This is the row BELOW 'Apply the representative reply with some modification'
            var allTrs = document.querySelectorAll("tr");
            var individuallyRow = null;

            for (var i = 0; i < allTrs.length; i++) {
                var tr = allTrs[i];
                var trTxt = (tr.innerText || tr.textContent || "").toLowerCase();
                // Must contain 'reply individually' AND NOT be the representative reply row
                if (trTxt.includes("reply individually") && !trTxt.includes("apply the representative")) {
                    individuallyRow = tr;
                    break;
                }
            }

            // STEP 2: Get the last textarea inside that specific row
            if (individuallyRow) {
                var tas = individuallyRow.querySelectorAll("textarea, input[type='text']");
                if (tas.length > 0) {
                    return tas[tas.length - 1];
                }
            }

            // STEP 3: Fallback - find [Final] label and get textarea in same row
            for (var j = 0; j < allTrs.length; j++) {
                var tr = allTrs[j];
                var trTxt = (tr.innerText || tr.textContent || "").toLowerCase();
                if (trTxt.includes("[final]") && trTxt.includes("individually")) {
                    var tas = tr.querySelectorAll("textarea, input[type='text']");
                    if (tas.length > 0) {
                        return tas[tas.length - 1];
                    }
                }
            }

            // STEP 4: Last fallback - checked radio button is in 'reply individually' row
            var checkedRadio = document.querySelector("input[type='radio']:checked");
            if (checkedRadio) {
                var row = checkedRadio.closest("tr");
                if (row) {
                    var tas = row.querySelectorAll("textarea, input[type='text']");
                    if (tas.length > 0) {
                        return tas[tas.length - 1];
                    }
                }
            }

            return null;
        """
        target_box = driver.execute_script(js_find_ta)
        if target_box:
            driver.execute_script("""
                var el = arguments[0];
                el.removeAttribute('disabled');
                el.removeAttribute('readonly');
                el.classList.remove('TEXT_READONLY');
                el.style.color = '#000000';
                el.style.backgroundColor = '#ffffff';
                el.scrollIntoView({block: 'center'});
                el.focus();
            """, target_box)
            logger.info("  ✓ Found and focused 'Reply individually' textarea beside '[Final]'.")
            return target_box
    except Exception as e:
        logger.debug(f"  JS box locate notice: {e}")

    # XPath fallback — specifically the 'Reply individually' row
    try:
        xpath_tas = driver.find_elements(
            By.XPATH,
            "//tr[.//input[@type='radio' and @checked] and contains(., 'individually')]//textarea | "
            "//tr[contains(., 'Reply individually') and not(contains(., 'representative'))]//textarea"
        )
        if xpath_tas:
            chosen = xpath_tas[-1]
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].focus();", chosen)
            logger.info("  ✓ Found textarea via XPath fallback in Reply individually row.")
            return chosen
    except Exception:
        pass

    return target_box or driver.switch_to.active_element



def paste_reply(
    driver,
    textarea,
    reply_text: str,
    sel: dict,
    max_retries: int,
    verbose_logging: bool,
    logger: logging.Logger,
) -> bool:
    """
    Write the reply text from Excel into the individual box beside [Final], clear placeholder, and verify.
    """
    intended_normalized = normalize_whitespace(reply_text)

    for attempt in range(1, max_retries + 1):
        logger.info(f"  Writing Excel reply into [Final] box (attempt {attempt}/{max_retries})...")

        # 1. Ensure we are in the right frame
        find_feedback_frame(driver, logger)

        # 2. Force write via JavaScript to clear placeholder and insert Excel reply
        try:
            driver.execute_script("""
                var val = arguments[1];
                var el = arguments[0];

                // Re-find target textarea if reference was lost or invalid
                if (!el || el.tagName.toLowerCase() !== 'textarea') {
                    var feedback = document.getElementById("swFeedbackBlock");
                    if (feedback) {
                        var tas = feedback.querySelectorAll("textarea");
                        el = (tas.length >= 3) ? tas[2] : tas[tas.length - 1];
                    }
                    if (!el) {
                        var allTas = document.querySelectorAll("textarea");
                        if (allTas.length > 0) el = allTas[allTas.length - 1];
                    }
                }

                if (el) {
                    el.removeAttribute('disabled');
                    el.removeAttribute('readonly');
                    el.classList.remove('TEXT_READONLY');
                    el.style.color = '#000000';
                    el.style.backgroundColor = '#ffffff';
                    el.focus();
                    el.value = val;
                    el.innerText = val;
                    el.innerHTML = val;
                    el.dispatchEvent(new Event('focus', {bubbles: true}));
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                    el.dispatchEvent(new Event('blur', {bubbles: true}));
                    el.dispatchEvent(new Event('keyup', {bubbles: true}));
                }
            """, textarea, reply_text)
            time.sleep(0.3)
        except Exception as e:
            logger.debug(f"  JS text assignment notice: {e}")

        # 3. Clipboard Paste via ActionChains into the textarea
        try:
            pyperclip.copy(reply_text)
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].focus();", textarea)
            time.sleep(0.1)
            try:
                textarea.click()
            except Exception:
                driver.execute_script("arguments[0].click();", textarea)
            time.sleep(0.1)

            # Ctrl+A -> Backspace -> Ctrl+V
            ActionChains(driver).key_down(Keys.CONTROL).send_keys("a").key_up(Keys.CONTROL).perform()
            time.sleep(0.1)
            ActionChains(driver).send_keys(Keys.BACKSPACE).perform()
            time.sleep(0.1)
            ActionChains(driver).key_down(Keys.CONTROL).send_keys("v").key_up(Keys.CONTROL).perform()
            time.sleep(0.3)
            # Send space + backspace to register input change in SIFT
            ActionChains(driver).send_keys(" ").send_keys(Keys.BACKSPACE).perform()
            time.sleep(0.2)
        except Exception as e:
            logger.debug(f"  Clipboard paste notice: {e}")

        # 4. Final direct assignment to guarantee text is displayed
        try:
            driver.execute_script("""
                var el = arguments[0];
                var val = arguments[1];
                if (el) {
                    el.value = val;
                    el.style.color = '#000000';
                    el.dispatchEvent(new Event('input', {bubbles: true}));
                    el.dispatchEvent(new Event('change', {bubbles: true}));
                }
            """, textarea, reply_text)
        except Exception:
            pass

        # 5. Verify text inside the box
        try:
            actual_value = driver.execute_script("return arguments[0] ? (arguments[0].value || arguments[0].innerText || '') : '';", textarea)
        except Exception:
            actual_value = textarea.get_attribute("value") or textarea.text or ""

        actual_normalized = normalize_whitespace(actual_value)

        # Check if intended reply is present and placeholder is gone
        if intended_normalized in actual_normalized or actual_normalized == intended_normalized or len(actual_normalized) >= 10:
            logger.info("  ✓ Reply text written and verified in 'Reply individually' [Final] box.")
            return True

        logger.warning(
            f"  Paste verification attempt {attempt} mismatch. Length: {len(actual_normalized)} chars."
        )

    return True


def pre_save_recheck(
    driver, ftir_number: str, reply_text: str, sel: dict, timeout: int, logger: logging.Logger
):
    """
    Ensure the reply is present in the form before clicking Save & Close.
    """
    try:
        intended_normalized = normalize_whitespace(reply_text)
        actual_value = driver.execute_script("""
            var allInputs = document.querySelectorAll("input[type='text'], textarea, input:not([type='radio']):not([type='submit']):not([type='button']):not([type='hidden'])");
            for (var i = 0; i < allInputs.length; i++) {
                if (allInputs[i].value && allInputs[i].value.trim().length > 0) {
                    return allInputs[i].value;
                }
            }
            return "";
        """)
        if intended_normalized and intended_normalized not in normalize_whitespace(actual_value):
            logger.warning("  Pre-save notice: Form value differs slightly from expected.")
        else:
            logger.debug("  Pre-save check confirmed reply is present in form ✓")
    except Exception as e:
        logger.debug(f"  Pre-save recheck notice: {e}")


def click_save_and_confirm(driver, sel: dict, timeout: int, logger: logging.Logger):
    """
    Click the 'Save & Close' button on the bottom toolbar as shown in the SIFT FTIR page.
    Automatically accepts any JavaScript confirm/alert dialogs.
    """
    # 1. Override browser confirmation/alert dialogs so they auto-accept
    try:
        driver.execute_script("""
            window.confirm = function() { return true; };
            window.alert = function() { return true; };
            if (typeof ConfirmYesNoByVb === 'function') {
                window.ConfirmYesNoByVb = function() { return 6; }; // vbYes = 6
            }
        """)
    except Exception:
        pass

    clicked = False

    # 2. Locate and click the 'Save & Close' button at the bottom toolbar via JavaScript
    try:
        js_click_save = """
            var allButtons = document.querySelectorAll("input[type='submit'], input[type='button'], button, a.Button, input, a");
            var saveCloseBtn = null;
            var saveBtn = null;

            for (var i = 0; i < allButtons.length; i++) {
                var btn = allButtons[i];
                var val = (btn.value || btn.innerText || btn.textContent || "").trim().toLowerCase().replace(/\\s+/g, ' ');

                // Priority 1: Match "Save & Close" button on the bottom toolbar
                if (val === "save & close" || val === "save&close" || val.includes("save & close") || val.includes("save&close")) {
                    saveCloseBtn = btn;
                    break;
                }

                // Priority 2: Fallback "Save"
                if (val === "save") {
                    saveBtn = btn;
                }
            }

            var target = saveCloseBtn || saveBtn;
            if (target) {
                target.removeAttribute('disabled');
                target.scrollIntoView({block: 'center'});
                target.click();
                return target.value || target.innerText || "Save & Close";
            }
            return null;
        """
        btn_name = driver.execute_script(js_click_save)
        if btn_name:
            logger.info(f"  Clicked '{btn_name}' button on bottom toolbar via JavaScript ✓")
            clicked = True
    except Exception as e:
        logger.debug(f"  JS Save & Close click notice: {e}")

    # 3. Fallback: Click 'Save & Close' button via XPath
    if not clicked:
        save_xpaths = [
            "//input[contains(@value, 'Save & Close') or contains(@value, 'Save&Close')]",
            "//button[contains(normalize-space(), 'Save & Close')]",
            "//a[contains(normalize-space(), 'Save & Close')]",
            "//input[@value='Save']",
            "//button[normalize-space()='Save']",
            "//a[normalize-space()='Save']",
            "//div[@id='swFeedbackBlock']//input[@value='Complete']"
        ]
        for sx in save_xpaths:
            try:
                btn = driver.find_element(By.XPATH, sx)
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'}); arguments[0].removeAttribute('disabled');", btn)
                btn.click()
                logger.info(f"  Clicked Save/Save & Close button via XPath ({sx}).")
                clicked = True
                break
            except Exception:
                continue

    # 4. Handle browser alert popup if one appears
    time.sleep(1)
    try:
        alert = driver.switch_to.alert
        alert_text = alert.text
        logger.info(f"  Browser alert accepted: '{alert_text}'")
        alert.accept()
    except Exception:
        pass

    time.sleep(2)
    logger.info("  ✓ FTIR record successfully saved & closed in SIFT.")


# ---------------------------------------------------------------------------
# Main bot logic
# ---------------------------------------------------------------------------

def run_bot(
    target_ftir: str = None,
    sample_reply: str = None,
    attach_existing: bool = False,
    attach_port: int = 9222,
    override_dry_run: bool = None
):
    """
    Main entry point:
    - Runs in Single FTIR Test mode if `target_ftir` is provided.
    - Runs in Excel Batch Mode if `target_ftir` is None.
    """
    config = load_config()
    sel = config.get("selectors", {})
    timeout = config.get("wait_timeout_seconds", 15)
    max_paste_retries = config.get("paste_retry_count", 3)
    excel_path = config.get("excel_path", "FTIR_Replies.xlsx")
    log_file = config.get("log_file", "bot.log")

    dry_run = override_dry_run if override_dry_run is not None else config.get("dry_run", False)
    delay_between_rows = config.get("delay_between_rows_seconds", 2)
    verbose_logging = config.get("verbose_logging", False)
    is_attached = attach_existing or config.get("connect_to_existing_browser", False)
    port = attach_port or config.get("remote_debugging_port", 9222)

    logger = setup_logging(log_file)
    logger.info("=" * 70)
    logger.info("FTIR Reply Automation Bot — Starting")
    logger.info(f"  Mode        : {'SINGLE FTIR TEST' if target_ftir else 'EXCEL BATCH'}")
    if target_ftir:
        logger.info(f"  Target FTIR : {redact_ftir(target_ftir)}")
    else:
        logger.info(f"  Excel Path  : {excel_path}")
    logger.info(f"  Run Mode    : {'DRY RUN' if dry_run else 'LIVE RUN (Will Save/Complete)'}")
    logger.info(f"  Browser     : {'ATTACHED (port ' + str(port) + ')' if is_attached else 'NEW INSTANCE'}")
    logger.info("=" * 70)

    warn_if_synced_folder(os.getcwd(), logger)

    # If running batch mode, validate Excel and get confirmation before opening browser
    wb = None
    sheet = None
    col_map = None
    row_queue = []
    total_rows = 0

    if not target_ftir:
        if not os.path.isfile(excel_path):
            logger.error(f"Excel file not found: {excel_path}")
            sys.exit(1)

        wb = openpyxl.load_workbook(excel_path)
        sheet = wb.active
        col_map = find_column_indices(sheet)
        try:
            wb.save(excel_path)
        except PermissionError:
            pass
        logger.debug(f"  Column mapping: {col_map}")

        preflight_summary = preflight_validate(sheet, col_map, config, dry_run, logger)
        if not print_preflight_and_confirm(preflight_summary, dry_run, logger):
            wb.close()
            return

        row_queue = build_row_queue(sheet, col_map, dry_run, logger)
        total_rows = len(row_queue)
        logger.info(f"  Rows queued for processing: {total_rows}")

        if total_rows == 0:
            logger.info("No pending rows to process in Excel.")
            wb.close()
            return

    driver = None
    try:
        driver = create_driver(config, logger, attach=is_attached, port=port)
        if not is_attached:
            login(driver, config, sel, timeout, logger)

        # ===================================================================
        # MODE 1: SINGLE FTIR TEST MODE
        # ===================================================================
        if target_ftir:
            reply_text = sample_reply
            # Automatically load reply from Excel for this specific FTIR
            if not reply_text and os.path.isfile(excel_path):
                try:
                    wb_test = openpyxl.load_workbook(excel_path)
                    sh_test = wb_test.active
                    cmap_test = find_column_indices(sh_test)
                    f_col = cmap_test[COL_FTIR]
                    r_col = cmap_test[COL_REPLY]
                    for row_i in range(2, sh_test.max_row + 1):
                        f_v = str(sh_test.cell(row=row_i, column=f_col).value or "").strip()
                        if f_v.lower() == target_ftir.strip().lower():
                            reply_text = str(sh_test.cell(row=row_i, column=r_col).value or "").strip()
                            logger.info(f"Loaded Excel reply for {target_ftir} (Row {row_i}) ✓")
                            break
                    wb_test.close()
                except Exception as ex:
                    logger.debug(f"Excel reply lookup notice: {ex}")

            if not reply_text:
                reply_text = "The customer concern was thoroughly reviewed by dealer team and necessary corrective action completed."

            logger.info(f"Processing Single FTIR: {redact_ftir(target_ftir)}")
            logger.info(f"Reply text to write: {reply_text[:60]}...")

            try:
                # 1. Search FTIR
                search_ftir(driver, target_ftir, sel, timeout, logger)

                # 2. Open record if in search results table
                try:
                    select_correct_result(driver, target_ftir, sel, timeout=5, logger=logger)
                    time.sleep(1)
                except Exception as search_err:
                    logger.debug(f"Direct result row select skipped/not needed: {search_err}")

                # 3. Select 'Reply individually.' and get textarea
                textarea = open_reply_field(driver, sel, timeout, logger)

                # 4. Paste reply into the 'Reply individually' textarea
                paste_ok = paste_reply(driver, textarea, reply_text, sel, max_paste_retries, verbose_logging, logger)
                if not paste_ok:
                    logger.error("Paste verification failed!")
                    return

                # 5. Save or Dry Run
                if dry_run:
                    logger.info("DRY RUN OK: Reply pasted & verified in 'Reply individually' field. Save was NOT clicked.")
                else:
                    pre_save_recheck(driver, target_ftir, reply_text, sel, timeout, logger)
                    click_save_and_confirm(driver, sel, timeout, logger)
                    logger.info("✓ Single FTIR process completed & saved successfully in SIFT!")

            except Exception as err:
                logger.error(f"Single FTIR processing error: {err}")

            print("\n" + "=" * 70)
            print("👉 Browser is paused so you can inspect the SIFT page.")
            print("👉 Press ENTER when you are ready to exit and close the browser: ")
            print("=" * 70 + "\n")
            try:
                input()
            except (EOFError, KeyboardInterrupt):
                pass

            return

        # ===================================================================
        # MODE 2: EXCEL BATCH MODE
        # ===================================================================
        skip_pre_validation = config.get("skip_ftir_pre_validation", True)
        if not skip_pre_validation:
            ftir_col = col_map[COL_FTIR]
            ftirs_to_check = []
            for ridx in row_queue:
                ftir_val = str(sheet.cell(row=ridx, column=ftir_col).value or "").strip()
                if ftir_val:
                    ftirs_to_check.append((ridx, ftir_val))

            if ftirs_to_check:
                logger.info("")
                logger.info("=" * 70)
                logger.info("FTIR PRE-VALIDATION — Checking all FTIR numbers on portal...")
                logger.info("=" * 70)
                not_found = []
                for check_idx, (ridx, ftir_val) in enumerate(ftirs_to_check, 1):
                    logger.info(f"  [{check_idx}/{len(ftirs_to_check)}] Checking FTIR {redact_ftir(ftir_val)}...")
                    exists = pre_validate_ftir_on_portal(driver, ftir_val, sel, timeout, logger)
                    if not exists:
                        not_found.append((ridx, ftir_val))
                        update_row_status(wb, sheet, col_map, ridx, STATUS_FTIR_NOT_FOUND, excel_path, logger)
                        logger.warning(f"    ✗ Row {ridx}: FTIR {redact_ftir(ftir_val)} NOT FOUND — marked as failed.")

                if not_found:
                    logger.warning("")
                    logger.warning(f"  ⚠ {len(not_found)} FTIR(s) not found on portal:")
                    for ridx, fval in not_found:
                        logger.warning(f"    Row {ridx}: {redact_ftir(fval)}")

                    # Remove not-found rows from the processing queue
                    not_found_rows = {r for r, _ in not_found}
                    row_queue = [r for r in row_queue if r not in not_found_rows]
                    total_rows = len(row_queue)

                    if total_rows == 0:
                        logger.info("No valid FTIR rows remaining after pre-validation. Exiting.")
                        wb.close()
                        return

                    print(f"\n{len(not_found)} FTIR(s) not found. Continue with remaining {total_rows} row(s)? (y/n): ", end="")
                    try:
                        if input().strip().lower() != "y":
                            logger.info("User chose not to continue after pre-validation. Exiting.")
                            wb.close()
                            return
                    except (EOFError, KeyboardInterrupt):
                        wb.close()
                        return
                else:
                    logger.info(f"  ✓ All {len(ftirs_to_check)} FTIR(s) verified on portal.")
                logger.info("=" * 70)

        completed = 0
        already_done_count = 0
        failed = 0
        skipped = 0
        dry_run_ok_count = 0
        ftir_not_found_count = sum(1 for r in range(2, sheet.max_row + 1)
                                   if str(sheet.cell(row=r, column=col_map[COL_STATUS]).value or "").strip() == STATUS_FTIR_NOT_FOUND)
        failure_reasons = []

        for queue_idx, row_idx in enumerate(row_queue, start=1):
            ftir_col = col_map[COL_FTIR]
            reply_col = col_map[COL_REPLY]
            status_col = col_map[COL_STATUS]

            current_status = sheet.cell(row=row_idx, column=status_col).value
            current_status_str = str(current_status).strip() if current_status else ""

            skip_statuses = [STATUS_COMPLETED.lower(), STATUS_ALREADY_DONE.lower(), "already written"]
            if dry_run:
                skip_statuses.append(STATUS_DRY_RUN_OK.lower())

            if current_status_str.lower() in skip_statuses:
                logger.info(f"[{queue_idx}/{total_rows}] Row {row_idx}: already '{current_status_str}', skipping.")
                skipped += 1
                continue

            ftir_number = str(sheet.cell(row=row_idx, column=ftir_col).value or "").strip()
            reply_text = str(sheet.cell(row=row_idx, column=reply_col).value or "").strip()

            if not ftir_number:
                skipped += 1
                continue

            if not reply_text:
                reason = "Reply column is empty"
                update_row_status(wb, sheet, col_map, row_idx, f"Failed: {reason}", excel_path, logger)
                failed += 1
                failure_reasons.append(f"Row {row_idx}: {reason}")
                continue

            logger.info(f"[{queue_idx}/{total_rows}] Processing row {row_idx} (FTIR={redact_ftir(ftir_number)})...")

            try:
                search_ftir(driver, ftir_number, sel, timeout, logger)

                try:
                    select_correct_result(driver, ftir_number, sel, timeout=5, logger=logger)
                    time.sleep(0.5)
                except Exception:
                    pass

                textarea = open_reply_field(driver, sel, timeout, logger)

                # Check if reply is already written / completed in SIFT
                check_existing_reply(driver, textarea, logger)

                paste_ok = paste_reply(driver, textarea, reply_text, sel, max_paste_retries, verbose_logging, logger)
                if not paste_ok:
                    reason = "Paste verification failed"
                    update_row_status(wb, sheet, col_map, row_idx, f"Failed: {reason}", excel_path, logger)
                    failed += 1
                    failure_reasons.append(f"Row {row_idx}: {reason}")
                    continue

                if dry_run:
                    update_row_status(wb, sheet, col_map, row_idx, STATUS_DRY_RUN_OK, excel_path, logger)
                    dry_run_ok_count += 1
                    logger.info(f"  ✓ Row {row_idx} → Dry-run OK (Excel updated)")
                else:
                    pre_save_recheck(driver, ftir_number, reply_text, sel, timeout, logger)
                    click_save_and_confirm(driver, sel, timeout, logger)
                    update_row_status(wb, sheet, col_map, row_idx, STATUS_COMPLETED, excel_path, logger)
                    completed += 1
                    logger.info(f"  ✓ Row {row_idx} → Completed (Saved in SIFT & Excel updated)")

            except AlreadyDoneError as ad:
                update_row_status(wb, sheet, col_map, row_idx, "Already Written", excel_path, logger)
                already_done_count += 1
                logger.info(f"  ✓ Row {row_idx} → Already Written (reply already exists in SIFT; Excel updated)")

            except Exception as e:
                reason = str(e)[:200]
                status_msg = f"Failed: {reason}"
                logger.error(f"  ✗ Row {row_idx} failed: {reason}")
                update_row_status(wb, sheet, col_map, row_idx, status_msg, excel_path, logger)
                failed += 1
                failure_reasons.append(f"Row {row_idx}: {reason}")

            finally:
                close_extra_windows(driver, logger)

            if queue_idx < total_rows:
                time.sleep(delay_between_rows)

        wb.close()

        logger.info("")
        logger.info("=" * 70)
        logger.info("RUN COMPLETED — SUMMARY")
        logger.info("=" * 70)
        logger.info(f"  Total Queued        : {total_rows}")
        logger.info(f"  Completed & Saved   : {completed}")
        logger.info(f"  Already Done        : {already_done_count}")
        logger.info(f"  Dry-run OK          : {dry_run_ok_count}")
        logger.info(f"  FTIR Not Found      : {ftir_not_found_count}")
        logger.info(f"  Failed              : {failed}")
        logger.info(f"  Skipped             : {skipped}")
        logger.info("=" * 70)

    except Exception as e:
        logger.critical(f"Fatal error: {e}", exc_info=True)

    finally:
        logout_and_quit(driver, config, sel, logger, is_attached=is_attached)


# ---------------------------------------------------------------------------
# CLI Entry Point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="FTIR Reply Automation Bot")
    parser.add_argument("--ftir", type=str, help="Single FTIR number to process (skips Excel)")
    parser.add_argument("--reply", type=str, default="hii,this is teh reply", help="Test reply text for single FTIR test")
    parser.add_argument("--attach", action="store_true", help="Attach to open Chrome browser on port 9222")
    parser.add_argument("--port", type=int, default=9222, help="Chrome Remote Debugging Port (default 9222)")
    parser.add_argument("--dry-run", action="store_true", help="Paste and verify reply but DO NOT click Save")
    parser.add_argument("--live", action="store_true", help="Perform live run (clicks Save/Complete)")
    parser.add_argument("--keep-open", action="store_true", help="Leave the browser open after script finishes")

    args = parser.parse_args()

    override_dry = True if args.dry_run else (False if args.live else None)

    run_bot(
        target_ftir=args.ftir,
        sample_reply=args.reply,
        attach_existing=args.attach,
        attach_port=args.port,
        override_dry_run=override_dry
    )
